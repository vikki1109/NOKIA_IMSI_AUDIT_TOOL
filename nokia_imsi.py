import os 
import sys 
import pandas as pd
import numpy as np 
import zipfile
import re
from openpyxl import load_workbook
import openpyxl
import xlsxwriter
from pandas import ExcelWriter

script_dir = os.path.abspath('')
input_dir=script_dir +"\input"
wb = openpyxl.Workbook() 
file_name = os.path.join(script_dir , "IMSI_OUTPUT.xlsx")




imsi_list = pd.read_csv(os.path.join(script_dir,"imsi_list.csv"))
wb.save(filename=file_name)
workbook = xlsxwriter.Workbook(file_name)
headers1=['NODE NAME','IMSI','CAMEL','PLMN NAME','VISITOR/HOME ','GT','CIPHERING','COUNTRY CODE LENGTH','MSRN GROUP','PNS TIME LIMIT','MSRN LIFE TIME','EXACT MS CATEGORY USAGE','REJECT CAUSE FOR UDL REJECTION','SUPPORT OF BOR','MOBILE TERMINATING ROAMING FORWARDING','ZONE CODES FROM HLR','REGIONAL ROAMING',' TLOC UP NEW VIS:','TLOC','TPER','TMO CALL','TMO SMS','TMT CALL','TMT SMS','TMT LOC REQ','TMT USSD','TMT SS OPER','TCS FALLBACK','ALOCVIS','ALOC','APER','AMO CALL','AMO SMS','AMT CALL','AMT SMS','AMT LOC REQ','AMT USSD','AMT SS OPER','ACS FALLBACK','ILOCVIS','ILOC','IPER','IMO CALL','IMO SMS','IMT CALL','IMT SMS','IMT LOC REQ','IMT USSD','IMT SS OPER','ICS FALLBACK','Black List Effect','Grey List Effect','Unknown IMEI CHECK','No Response effect']
worksheet_sheet1 = workbook.add_worksheet('sheet1')

I = 0
for header1 in headers1:
    worksheet_sheet1.write(0, I, header1)
    
    I = I + 1



def WriteDeviationData(worksheet,ROW,data_row):
    import pdb
    error = []
    try:
        I=0
        for data in data_row:
            worksheet.write(ROW,I,data)
            I=I+1
    except Exception as e:
        error.append("Error: Write Deviation Data in Excel File")
        if not str(e).find('nan_inf_to_errors') >=0:
            print("WriteDeviationData:"+str(e))
    finally:
        return error

#with open(os.path.join(script_dir,'kolkata imsi logs complete output'),'r') as f:
imsis1={}
plmn_name={}
plmn_array={}
gt1={}
for file in os.listdir(input_dir):_
    with open(os.path.join(input_dir,file),encoding="ISO-8859-1") as f:
        lines = f.readlines()
        new_lines = []
        for line in lines:
            if (len(line.strip()) > 0):
                new_lines.append(line.strip())
            
            if 'COMMAND EXECUTED' in line:
                break
        lines = new_lines

        flag =0
        
        for i in range(2,len(lines)):
            if('IMSI            IMSI PLMN            GT              NP    TON NI  SPC  (SPCDEC)' in lines[i]):
                v=0``
                for j in range(i+2,len(lines)):
                    # import pdb; pdb.set_trace()
                    if (j != (len(lines)-1)):
                        if lines[j] != " ":
                            plmn_name[j]=(lines[j])
                            temp_name=plmn_name[j]
                            plmn_array[v]=str(temp_name[19:36]).strip()
                            imsis1[v]=str(temp_name[0:16]).strip()
                            gt1[v]=str(temp_name[37:44]).strip()
                            v=v+1
    if ('COMMAND EXECUTED' in lines[i]):
        break
plmn_ar={}
gt2={}
x=1
imsi1={}
for imsi in imsi_list['IMSI']:
    for z in range (len(plmn_array)):
        if str(imsi).strip() == imsis1[z]:
                plmn_ar[imsi]=plmn_array[z]
                
                gt2[imsi]=gt1[z]
                imsi1[imsi]=imsi
                # import pdb; pdb.set_trace()
                break

            # print(imsis1[z])
            # print(plmn_array[z])
            # print(gt1[z])

                
                

            # print(imsis1[z])
            # print(plmn_array[z])
            # print(gt1[z])

print(plmn_ar.items())

cond=True
for file in os.listdir(input_dir):
    #import pdb; pdb.set_trace()
    node_name,imsis1,gt,imsis,hpl,vpl,camel,camel1,vplmn_name,hplmn_name,gt,ciphering,country_code_length,country_code_length1,msrn_group,msrn_group1,pns_time_limit,pns_time_limit1,msrn_life_time,msrn_life_time1,exact_ms_category_usage,reject_cause_for_udl_rejection,support_of_bor,mobile_terminating_roaming_forwarding,zone_codes_from_hlr,regional_roaming,tlocvis,tlocvis1,tloc,tloc1,tper,tmoc,tmoc1, tmos, tmtc,tmtc1, tmts,tmts1,tmtl,tmtu,tmtss,tmcs,amtl,amtu,amtss,amcs,imtl,imtu,imtss,imcs, alocvis,alocvis1,aloc,aloc1, aper,amoc,amoc1,amos, amtc,amtc1,amts,amts1,ilocvis,ilocvis1,iloc,iloc1, iper,imoc,imoc1, imos,imtc,imtc1, imts,imts1,ble,mtrf,gle,uie,nre=['']*79   
    if ('nokia_imsi' not in file) and ('imsi_parse' not in file) and ('IMSI_OUTPUT' not in file):
        with open(os.path.join(input_dir,file),encoding="ISO-8859-1") as f:
            content = f.readlines()
            content = [x.strip() for x in content]
            try:
                A=1
                for index , line in enumerate(content):
                    for key,value in plmn_ar.items():
        
                        if 'MSS       ' in line:
                            if  "VISITOR PLMN "+ value+" " in content[index+4] or "HOME PLMN "+ value + " " in content[index+4] :
                                # import pdb; pdb.set_trace()
                                B=1
                                node_name= line.split('MSS       ')[1].strip().replace(';','')
                                node_name1= node_name.split()[0]
                                vpl=line.split()[0]
                            #print(node_name1)  
                                imsis,camel,camel1,vpl,hpl,gt,ciphering,country_code_length,msrn_group,pns_time_limit,msrn_life_time,exact_ms_category_usage,reject_cause_for_udl_rejection,support_of_bor,mobile_terminating_roaming_forwarding,zone_codes_from_hlr,regional_roaming,tlocvis,tlocvis1,tloc,tloc1,tper,tmoc,tmoc1, tmos, tmtc,tmtc1, tmts,tmts1,tmtl,tmtu,tmtss,tmcs,amtl,amtu,amtss,amcs,imtl,imtu,imtss,imcs,alocvis,alocvis1, aloc,aloc1, aper,amoc,amoc1,amos, amtc,amtc1,amts,amts1,ilocvis,ilocvis1,iloc,iloc1, iper,imoc,imoc1, imos,imtc,imtc1, imts,imts1,ble,mtrf,gle,uie,nre=['']*70
                            
                                hplmn_name=plmn_ar[key]
                                gt=gt2[key]
                                imsi=key
                                for ind2,line2 in enumerate(content[index+5:]):
                                    if 'VISITOR' in line2:
                                        vpl= 'VISITOR'
                                    if 'HOME' in line2:
                                        hpl='HOME'
                                    if 'SUPPORTED CAMEL PHASE:' in line2:
                                        camel = line2.split('SUPPORTED CAMEL PHASE:')[1] 
                                        camel1=camel[13:26]                                           
                                    if 'CIPHERING:     ' in line2:
                                        ciphering=line2.split('CIPHERING:')[1].strip()
                                    if 'COUNTRY CODE LENGTH: ' in line2:
                                        country_code_length=line2.split('COUNTRY CODE LENGTH: ')[1]
                                        country_code_length1=country_code_length.split('2')[0]
                                    if 'MSRN GROUP:' in line2:
                                        msrn_group=line2.split('MSRN GROUP:')[1]
                                        msrn_group1=msrn_group.split()[0]
                                    if 'MSRN LIFE TIME:' in line2:
                                        msrn_life_time=line2.split('MSRN LIFE TIME:')[1]
                                        msrn_life_time1=msrn_life_time.split()[0].strip()
                                    if 'PNS TIME LIMIT:' in line2:
                                        pns_time_limit=line2.split('PNS TIME LIMIT:')[1]
                                        pns_time_limit1=pns_time_limit.split()[0].strip()
                                    if 'EXACT MS CATEGORY USAGE:                'in line2:
                                        exact_ms_category_usage=line2.split('EXACT MS CATEGORY USAGE:                ')[1].strip()
                                    if 'REJECT CAUSE FOR UDL REJECTION:         ' in line2:
                                        reject_cause_for_udl_rejection=line2.split('REJECT CAUSE FOR UDL REJECTION:         ')[1].strip()
                                    if 'SUPPORT OF BOR: ' in line2:
                                        support_of_bor=line2.split('SUPPORT OF BOR: ')[1].strip()
                                    if 'MOBILE TERMINATING ROAMING FORWARDING:' in line2:
                                        mobile_terminating_roaming_forwarding=line2.split('MOBILE TERMINATING ROAMING FORWARDING:')[1].strip()
                                        mobile_terminating_roaming_forwarding.replace(' ', '')
                                    if 'ZONE CODES FROM HLR:' in line2:
                                        zone_codes_from_hlr=line2.split('ZONE CODES FROM HLR:')[1].strip()
                                    if 'REGIONAL ROAMING:                       ' in line2:
                                        regional_roaming=line2.split('REGIONAL ROAMING:                       ')[1]
                                    if 'GREY LIST EFFECT: ' in line2:
                                        gle=line2.split('GREY LIST EFFECT: ')[1].strip()
                                    

                                    #if 'TMSI ALLOCATION ' in line2 :
                                    if 'LOC UP NEW VIS:' in line2:
                                        tlocvis = line2.split('LOC UP NEW VIS:')[1]
                                        tlocvis1=tlocvis.split()[0]
                                        # tlocvis=tlocvis.split()[0]
                                        # import pdb; pdb.set_trace()
                                    if 'LOC UP: ' in line2:
                                        tloc = line2.split('LOC UP:')[1]
                                        tloc1=tloc.split()[0]
                                    if 'PER UP:' in line2:
                                        tper='5'
                                    if 'MO CALL:  ' in line2 :
                                        tmoc=line2.split('MO CALL:  ')[1]
                                        tmoc1=tmoc.split()[0]
                                    if 'MO SMS: ' in line2:
                                        tmos=line2.split('MO SMS:  ')[1]
                                    if 'MT CALL: ' in line2:
                                        tmtc=line2.split('MT CALL:')[1]
                                        tmtc1=tmtc.split()[0]
                                    if 'MT SMS:' in line2:
                                        tmts=line2.split('MT SMS:')[1]
                                        tmts1='10'
                                    if 'MT LOC REQ:' in line2:
                                        tmts=line2.split('MT LOC REQ:')[1]
                                        tmts1='10'
                                    if 'MT SMS:' in line2:
                                        tmts=line2.split('MT SMS:')[1]
                                        tmts1='10'
                                    if 'MT LOC REQ:   ' in line2:
                                        tmtl=line2.split('MT LOC REQ:   ')[1]
                                        tmtl=tmtl.split()[0]
                                    
                                    if 'MT USSD:         ' in line2:
                                        tmtu=line2.split('MT USSD:         ')[1]
                                        tmtu=tmtu.split()[0]
                         
                                    if 'SS OPER:  ' in line2:
                                        tmtss=line2.split('SS OPER:  ')[1]
                                        tmtss=tmtss.split()[0]
                              
                                    if 'CS FALLBACK:' in line2:
                                        tmcs=line2.split('CS FALLBACK:')[1]
                                        tmcs='0'
                            
                                    
                                        #if 'AUTHENTICATION ' in line2 :
                                    if 'LOC UP NEW VIS:' in line2:
                                        alocvis = line2.split('LOC UP NEW VIS:')[1]
                                        alocvis1=alocvis.split()[0]
                                    if 'LOC UP: ' in line2:
                                        aloc = line2.split('LOC UP:')[1]
                                        aloc1=aloc.split()[0]
                                    if 'PER UP:' in line2:
                                        aper='5'
                                    if 'MO CALL:  ' in line2 :
                                        amoc=line2.split('MO CALL:  ')[1]
                                        amoc1=amoc.split()[0]
                                    if 'MO SMS: ' in line2:
                                        amos=line2.split('MO SMS:  ')[1]
                                    if 'MT CALL: ' in line2:
                                        amtc=line2.split('MT CALL: ')[1]
                                        amtc1=amtc.split()[0]
                                    if 'MT SMS:' in line2:
                                        amts=line2.split('MT SMS:')[1]
                                        amts1='10'
                                    if 'MT LOC REQ:   ' in line2:
                                        amtl=line2.split('MT LOC REQ:   ')[1]
                                        amtl=amtl.split()[0]
                              
                                    if 'MT USSD:         ' in line2:
                                        amtu=line2.split('MT USSD:         ')[1]
                                        amtu=amtu.split()[0]
                                
                                    if 'SS OPER:  ' in line2:
                                        amtss=line2.split('SS OPER:  ')[1]
                                        amtss=amtss.split()[0]
                                  
                                    if 'CS FALLBACK:' in line2:
                                        amcs=line2.split('CS FALLBACK:')[1]
                                        amcs='0'
                                 
                                #  if 'IMEI CHECKING' in line2 :
                                    if 'LOC UP NEW VIS:' in line2:
                                        ilocvis = line2.split('LOC UP NEW VIS:')[1]
                                        ilocvis1=ilocvis.split()[0]
                                    if 'LOC UP: ' in line2:
                                        iloc = line2.split('LOC UP:')[1]
                                        iloc1=iloc.split()[0]
                                    if 'PER UP:' in line2:
                                        iper=line2.split('PER UP:')[1]
                                    if 'MO CALL:  ' in line2 :
                                        imoc=line2.split('MO CALL:  ')[1]
                                        imoc1=imoc.split()[0]
                                    if 'MO SMS: ' in line2:
                                        imos=line2.split('MO SMS:  ')[1]
                                    if 'MT CALL:' in line2:
                                        imtc=line2.split('MT CALL:')[1]
                                        imtc1=imtc.split()[0]
                                    if 'MT SMS:' in line2:
                                        imts=line2.split('MT SMS:')[1]
                                        imts1=imts.split()[0]
                                    if 'MT LOC REQ:   ' in line2:
                                        imtl=line2.split('MT LOC REQ:   ')[1]
                                        imtl=imtl.split()[0]
                         
                                    if 'MT USSD:         ' in line2:
                                        imtu=line2.split('MT USSD:         ')[1]
                                        imtu=imtu.split()[0]
                                   
                                    if 'SS OPER:  ' in line2:
                                        imtss=line2.split('SS OPER:  ')[1]
                                        imtss=imtss.split()[0]
                                   
                                    if 'CS FALLBACK:' in line2:
                                        imcs=line2.split('CS FALLBACK:')[1]
                                    if 'BLACK LIST EFFECT:'in line2:
                                        ble=line2.split('BLACK LIST EFFECT:')[1].strip()
                                    if 'UNKNOWN IMEI EFFECT: ' in line2:
                                        uie=line2.split('UNKNOWN IMEI EFFECT: ')[1].strip()
                                    if '  NO RESPONSE EFFECT: ' in line2:
                                        nre=line2.split('  NO RESPONSE EFFECT: ')[1].strip()
    
                                    if 'IMS CENTRALIZED SERVICES PARAMETERS'  in line2: 
                                        # import pdb; pdb.set_trace()
                                        break
                                    if 'ZQNS;' in line2:
                                        break 
       
                                    
                                
                                to_put=[node_name1,imsi,camel1,hplmn_name,vpl,gt,ciphering,country_code_length1,msrn_group1,pns_time_limit1,msrn_life_time1,exact_ms_category_usage,reject_cause_for_udl_rejection,support_of_bor,mobile_terminating_roaming_forwarding,zone_codes_from_hlr,regional_roaming,tlocvis1,tloc1,tper,tmoc1, tmos, tmtc1, tmts1,tmtl,tmtu,tmtss,tmcs,alocvis1, aloc1, aper,amoc1,amos, amtc1,amts1,amtl,amtu,amtss,amcs,ilocvis1,iloc1, iper,imoc1, imos,imtc1, imts1,imtl,imtu,imtss,imcs,ble,gle,uie,nre]
                                
                                WriteDeviationData(worksheet_sheet1, A, to_put)
                                
                                A = A + 1
                            
                            
                                     
                                    
                                    
                        
                                
            except Exception as e:
                
                print("Exception found as ", str(e))
                import traceback
                print(traceback.format_exc())
workbook.close()
